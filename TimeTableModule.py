from SubjectEntryModule import *
from openpyxl import *
from openpyxl.styles import *

class TimeTable:
    def __init__(self, subjectList, step=30):
        self.subjectList = subjectList
        self.step = step
        self.baseTimeRange = TimeRange.findTimeRange(subjectList)

    def rowsFromTimeRange(self, timeRange):
        firstRow = 2 + (Time.distance(self.baseTimeRange.initial, timeRange.initial) // self.step)
        rows = [i + firstRow for i in range(len(timeRange)//self.step)]
        return rows

    def getCells(self, day, timeRange):
        column = day.getColumn()
        rows = self.rowsFromTimeRange(timeRange)

        cells = [column+str(row) for row in rows]
        return cells

    def addSubject(self, ws, subject):

        for day in subject.days:
            cells = self.getCells(day, subject.timeRange)
            ws.merge_cells("{}:{}".format(cells[0], cells[-1]))
            for cell in cells:
                ws[cell] = "{}\n{}".format(subject.name, str(subject.module))
                ws[cell].style = "subjectStyle"

    def generateTable(self, filename):
        wb = Workbook()
        ws = wb.active
        self.initStyle(wb)
        self.addDays(ws)
        self.addTimeMarks(ws)

        for subject in self.subjectList:
            self.addSubject(ws, subject)

        #self.setTableBorder(ws)
        wb.save(filename)

    def setTableBorder(self, ws):
        thickSide = Side(style="thick", color="000000")
        boldBorder = Border(right=thickSide, left=thickSide, top=thickSide, bottom=thickSide)

        cellsRange = "A1:F{}".format(1+len(self.baseTimeRange)//self.step)
        rows = ws[cellsRange]

        for row in rows:
            row[0].border = Border(left=thickSide)
            row[-1].border = Border(right=thickSide)
        for cell in rows[0]:
            cell.border = Border(top=thickSide)

    def addDays(self, ws):
        ws["B1"] = "Lunes"
        ws["C1"] = "Martes"
        ws["D1"] = "Miércoles"
        ws["E1"] = "Jueves"
        ws["F1"] = "Viernes"

        for colName in ["A","B","C","D","E","F"]:
            col = ws.column_dimensions[colName]
            col.width = 18

    def addTimeMarks(self, ws):
        timeList = TimeRange.timeRangeArray(self.baseTimeRange, self.step)
        i = 0
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=len(timeList)):
            for cell in row:
                cell.value = str(TimeRange(timeList[i],timeList[i+1]))
                i+=1

    def initStyle(self, wb):
        centered = Alignment(horizontal="center", vertical="center", wrapText=True)
        mediumSide = Side(style="medium", color="000000")
        border = Border(left=mediumSide, right=mediumSide, top=mediumSide, bottom=mediumSide)
        subjectStyle = NamedStyle(name="subjectStyle", border=border, alignment=centered)
        wb.add_named_style(subjectStyle)
